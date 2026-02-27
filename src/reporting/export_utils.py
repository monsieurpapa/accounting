"""
PDF and Excel export utilities for reporting views.

Uses reportlab for PDF and openpyxl for Excel (xlsx) generation.
"""
from io import BytesIO
from decimal import Decimal
from datetime import date

from django.http import HttpResponse


def _format_decimal(val):
    """Format Decimal or number for display."""
    if val is None:
        return "0.00"
    if isinstance(val, Decimal):
        return f"{float(val):,.2f}"
    return f"{float(val):,.2f}"


def export_pdf_general_ledger(entry_lines, selected_account, generation_date, fiscal_year=None, period=None):
    """Generate PDF for general ledger report."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    elements = []

    title = Paragraph("General Ledger / Grand Livre", styles['Heading1'])
    elements.append(title)
    elements.append(Spacer(1, 0.5*cm))

    if selected_account:
        elements.append(Paragraph(f"<b>Account:</b> {selected_account.code} - {selected_account.name}", styles['Normal']))
    if fiscal_year:
        elements.append(Paragraph(f"<b>Fiscal Year:</b> {fiscal_year.name}", styles['Normal']))
    if period:
        elements.append(Paragraph(f"<b>Period:</b> {period.name}", styles['Normal']))
    elements.append(Paragraph(f"<b>Generated:</b> {generation_date}", styles['Normal']))
    elements.append(Spacer(1, 0.5*cm))

    headers = ["Date", "Reference", "Description", "Debit", "Credit"]
    data = [headers]
    for line in entry_lines:
        je = line.journal_entry
        data.append([
            str(je.date),
            je.reference or "-",
            line.description or "-",
            _format_decimal(line.debit_amount),
            _format_decimal(line.credit_amount),
        ])

    if len(data) == 1:
        data.append(["No entries", "-", "-", "-", "-"])

    t = Table(data, colWidths=[3*cm, 3*cm, 8*cm, 3*cm, 3*cm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (3, 0), (4, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
    ]))
    elements.append(t)

    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


def export_excel_general_ledger(entry_lines, selected_account, generation_date, fiscal_year=None, period=None):
    """Generate Excel for general ledger report."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "General Ledger"

    ws.append(["General Ledger / Grand Livre"])
    ws.merge_cells('A1:E1')
    ws['A1'].font = Font(bold=True, size=14)
    if selected_account:
        ws.append([f"Account: {selected_account.code} - {selected_account.name}"])
    if fiscal_year:
        ws.append([f"Fiscal Year: {fiscal_year.name}"])
    if period:
        ws.append([f"Period: {period.name}"])
    ws.append([f"Generated: {generation_date}"])
    ws.append([])

    headers = ["Date", "Reference", "Description", "Debit", "Credit"]
    ws.append(headers)
    for row in headers:
        ws.cell(row=ws.max_row, column=headers.index(row) + 1).font = Font(bold=True)

    for line in entry_lines:
        je = line.journal_entry
        ws.append([
            str(je.date),
            je.reference or "-",
            line.description or "-",
            float(line.debit_amount) if line.debit_amount else 0,
            float(line.credit_amount) if line.credit_amount else 0,
        ])

    for col in range(1, 6):
        ws.column_dimensions[chr(64 + col)].width = 15
    ws.column_dimensions['C'].width = 30

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def export_pdf_balance_sheet(assets, liabilities, equity, total_assets, total_liabilities, total_equity, as_of_date, generation_date):
    """Generate PDF for balance sheet."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("Balance Sheet / Bilan", styles['Heading1']))
    elements.append(Paragraph(f"As of: {as_of_date}", styles['Normal']))
    elements.append(Paragraph(f"Generated: {generation_date}", styles['Normal']))
    elements.append(Spacer(1, 0.5*cm))

    data = [["Code", "Name", "Balance"]]
    for row in assets:
        data.append([row['code'], row['name'], _format_decimal(row['balance'])])
    data.append(["", "TOTAL ASSETS", _format_decimal(total_assets)])
    for row in liabilities:
        data.append([row['code'], row['name'], _format_decimal(row['balance'])])
    data.append(["", "TOTAL LIABILITIES", _format_decimal(total_liabilities)])
    for row in equity:
        data.append([row['code'], row['name'], _format_decimal(row['balance'])])
    data.append(["", "TOTAL EQUITY", _format_decimal(total_equity)])

    t = Table(data, colWidths=[3*cm, 10*cm, 4*cm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
    ]))
    elements.append(t)
    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


def export_excel_balance_sheet(assets, liabilities, equity, total_assets, total_liabilities, total_equity, as_of_date, generation_date):
    """Generate Excel for balance sheet."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Balance Sheet"
    ws.append(["Balance Sheet / Bilan"])
    ws.append([f"As of: {as_of_date}", f"Generated: {generation_date}"])
    ws.append([])
    ws.append(["Code", "Name", "Balance"])
    ws['A4'].font = ws['B4'].font = ws['C4'].font = Font(bold=True)
    for row in assets:
        ws.append([row['code'], row['name'], float(row['balance'])])
    ws.append(["", "TOTAL ASSETS", float(total_assets)])
    for row in liabilities:
        ws.append([row['code'], row['name'], float(row['balance'])])
    ws.append(["", "TOTAL LIABILITIES", float(total_liabilities)])
    for row in equity:
        ws.append([row['code'], row['name'], float(row['balance'])])
    ws.append(["", "TOTAL EQUITY", float(total_equity)])
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def export_pdf_income_statement(revenues, expenses, total_revenue, total_expense, net_income, start_date, end_date, generation_date):
    """Generate PDF for income statement."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("Income Statement / Compte de Resultat", styles['Heading1']))
    elements.append(Paragraph(f"Period: {start_date} to {end_date}", styles['Normal']))
    elements.append(Paragraph(f"Generated: {generation_date}", styles['Normal']))
    elements.append(Spacer(1, 0.5*cm))

    data = [["Code", "Name", "Amount"]]
    for row in revenues:
        data.append([row['code'], row['name'], _format_decimal(row['balance'])])
    data.append(["", "TOTAL REVENUE", _format_decimal(total_revenue)])
    for row in expenses:
        data.append([row['code'], row['name'], _format_decimal(row['balance'])])
    data.append(["", "TOTAL EXPENSES", _format_decimal(total_expense)])
    data.append(["", "NET INCOME", _format_decimal(net_income)])

    t = Table(data, colWidths=[3*cm, 10*cm, 4*cm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
    ]))
    elements.append(t)
    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


def export_excel_income_statement(revenues, expenses, total_revenue, total_expense, net_income, start_date, end_date, generation_date):
    """Generate Excel for income statement."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Income Statement"
    ws.append(["Income Statement / Compte de Resultat"])
    ws.append([f"Period: {start_date} to {end_date}", f"Generated: {generation_date}"])
    ws.append([])
    ws.append(["Code", "Name", "Amount"])
    ws['A4'].font = ws['B4'].font = ws['C4'].font = Font(bold=True)
    for row in revenues:
        ws.append([row['code'], row['name'], float(row['balance'])])
    ws.append(["", "TOTAL REVENUE", float(total_revenue)])
    for row in expenses:
        ws.append([row['code'], row['name'], float(row['balance'])])
    ws.append(["", "TOTAL EXPENSES", float(total_expense)])
    ws.append(["", "NET INCOME", float(net_income)])
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def export_pdf_trial_balance(report_data, grand_totals, fiscal_year, period, start_date, end_date, generation_date):
    """Generate PDF for trial balance."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=1.5*cm, leftMargin=1.5*cm, topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    elements = []

    period_str = period.name if period else (fiscal_year.name if fiscal_year else f"{start_date} - {end_date}")
    elements.append(Paragraph("Trial Balance / Balance de Verification", styles['Heading1']))
    elements.append(Paragraph(f"Period: {period_str}", styles['Normal']))
    elements.append(Paragraph(f"Generated: {generation_date}", styles['Normal']))
    elements.append(Spacer(1, 0.3*cm))

    headers = ["Code", "Name", "Opening Debit", "Opening Credit", "Period Debit", "Period Credit", "Closing Debit", "Closing Credit"]
    data = [headers]
    for row in report_data:
        data.append([
            row['code'], row['name'],
            _format_decimal(row['opening_debit']), _format_decimal(row['opening_credit']),
            _format_decimal(row['period_debit']), _format_decimal(row['period_credit']),
            _format_decimal(row['closing_debit']), _format_decimal(row['closing_credit']),
        ])
    data.append([
        "", "TOTALS",
        _format_decimal(grand_totals['opening_debit']), _format_decimal(grand_totals['opening_credit']),
        _format_decimal(grand_totals['period_debit']), _format_decimal(grand_totals['period_credit']),
        _format_decimal(grand_totals['closing_debit']), _format_decimal(grand_totals['closing_credit']),
    ])

    col_widths = [2*cm, 5*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2.5*cm]
    t = Table(data, colWidths=col_widths)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
    ]))
    elements.append(t)
    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


def export_excel_trial_balance(report_data, grand_totals, fiscal_year, period, start_date, end_date, generation_date):
    """Generate Excel for trial balance."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Trial Balance"
    period_str = period.name if period else (fiscal_year.name if fiscal_year else f"{start_date} - {end_date}")
    ws.append(["Trial Balance / Balance de Verification"])
    ws.append([f"Period: {period_str}", f"Generated: {generation_date}"])
    ws.append([])
    headers = ["Code", "Name", "Opening Debit", "Opening Credit", "Period Debit", "Period Credit", "Closing Debit", "Closing Credit"]
    ws.append(headers)
    for c in range(1, 9):
        ws.cell(row=ws.max_row, column=c).font = Font(bold=True)
    for row in report_data:
        ws.append([
            row['code'], row['name'],
            float(row['opening_debit']), float(row['opening_credit']),
            float(row['period_debit']), float(row['period_credit']),
            float(row['closing_debit']), float(row['closing_credit']),
        ])
    ws.append([
        "", "TOTALS",
        float(grand_totals['opening_debit']), float(grand_totals['opening_credit']),
        float(grand_totals['period_debit']), float(grand_totals['period_credit']),
        float(grand_totals['closing_debit']), float(grand_totals['closing_credit']),
    ])
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
